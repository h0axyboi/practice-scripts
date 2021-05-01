print('''
Instructions for the 3x3x3 blindfolded mnemonics database:
Enter the two letters to see if there exists a mnemonic
f there isn't, then type 'new' for a new entry and enter the two letters and the mnemonic you want
Example: IL ILL
'exit' to stop
''')


from openpyxl import load_workbook
wb=load_workbook('mnemonics_database.xlsx')
ws=wb['Sheet1']
d={}
for row in range(2,ws.max_row+1):
    a=ws['A'+str(row)].value
    b=ws['B'+str(row)].value
    d[a]=b

while True:
    inp=input()
    if inp=='exit':
        break
    if inp=='new':
        entry=input()
        x=entry.index(' ')
        a=entry[:x].upper()
        b=entry[x+1:].upper()
        ws['A'+str(ws.max_row+1)]=a
        ws['B'+str(ws.max_row)]=b
        d[a]=b
    elif inp.upper() in d.keys():
        print(d[inp.upper()])
    else:
        print('No mnemonic exists')

wb.save('mnemonics_database.xlsx')
